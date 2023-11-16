"""[Save all xl Sheets as seprate sheets each sheet in 1 excel file]

    """

##
import pathlib
import warnings

import multiprocess as mp
import openpyxl as xl
import pandas as pd
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from _0 import get_all_fps_wt_suffix
from _1 import smpl_ret_cache_fp
from _2 import get_pre_mod_name
from _2 import save_current_module_cache
from _3 import count_notna_of_thecol_in_groups
from _3 import load_update_df
from _3 import return_index_level_names_from_to
from Helper.name_space import cc
from Helper.name_space import d
from Helper.name_space import indi

warnings.filterwarnings("ignore")

Module_name = '_6.py'

class ExcelWorkBook :
  def __init__(
      self ,
      xl_stem
      ) :
    self.stem = xl_stem
    self.fp0 = (d.xl / self.stem).with_suffix('.xlsx')
    self.sheetname_hidestat: dict = {}
    self.sheetname_isempty: dict = {}

  def _print(
      self ,
      *args
      ) :
    print(self.fp0.name , ': ' , *args)

  def _load_workbook(
      self
      ) :
    self.wb = xl.load_workbook(self.fp0 , data_only = True)

  def _make_sheetname_hiddenstat_dict(
      self
      ) :
    for _i , shn in enumerate(self.wb.sheetnames) :
      self.sheetname_hidestat[shn] = self.wb[shn].sheet_state == 'hidden'
      _df = pd.DataFrame(self.wb[shn].values)
      self.sheetname_isempty[shn] = _df.empty
      if not _df.empty :
        fp = d.xl_Sheets / f'{self.stem}_{_i}.xlsx'
        save_df_as_a_nice_xl(_df , fp , header = False)
    self._print('Sheets Saved.')

  def process(
      self
      ) :
    self._load_workbook()
    self._make_sheetname_hiddenstat_dict()
    return self.sheetname_hidestat , self.sheetname_isempty

def save_df_as_a_nice_xl(
    df: pd.DataFrame ,
    fpn: pathlib.Path ,
    index: bool = False ,
    header: bool = True ,
    max_col_length: int = 40
    ) :
  wb = pyxl.Workbook()
  ws = wb.active
  df = df.fillna(value = '')
  for r in dataframe_to_rows(df , index = index , header = header) :
    ws.append(r)
  panes = index * ws['A'] + header * ws[1]
  for cell in panes :
    cell.style = 'Pandas'
  for column in ws.columns :
    length = max(len(str(cell.value)) for cell in column) + 3
    length = length if length <= max_col_length else max_col_length
    ws.column_dimensions[column[0].column_letter].width = length
  ws.freeze_panes = 'A2'
  wb.save(fpn)
  wb.close()

def _targ(
    xl_stem
    ) :
  try :
    pdfrep = ExcelWorkBook(xl_stem)
    out = pdfrep.process()
    return out
  except (AttributeError , KeyError) as e :
    print(e)
    return {} , {}

def main() :
  pass

  ##

  pre_mod_name = get_pre_mod_name(d.code , Module_name)

  ##
  cur_mod_cache_fp = smpl_ret_cache_fp(Module_name)
  pre_mod_cache_fp = smpl_ret_cache_fp(pre_mod_name)

  ##
  new_cols = {
      indi.L3_SheetNum         : None ,
      cc.is_sheet_hidden       : None ,
      cc.sheet_name            : None ,
      cc.sheet_count           : None ,
      cc.not_hidden_sheets_cou : None ,
      cc.sheetStem             : None ,
      cc.is_empty              : None ,
      }

  ##
  df = load_update_df(pre_module_cache_fp = pre_mod_cache_fp ,
                      current_module_cache_fp = cur_mod_cache_fp ,
                      new_cols_2add_and_update = list(new_cols.keys()))

  ##
  drop_cols = {
      cc.dlLink        : None ,
      cc.attFilesCount : None ,
      cc.isFileHealthy : None ,
      }

  df = df.drop(columns = drop_cols , errors = 'ignore')

  ##
  def remove_excess_sheets(
      idf
      ) :
    fps = get_all_fps_wt_suffix(d.xl_Sheets , '.xlsx')
    sts = [x.stem for x in fps]
    print(len(sts))

    sht_2del = set(sts) - set(idf[cc.sheetStem])
    print('Sheets to remove count:' , len(sht_2del))

    _ = [(d.xl_Sheets / x).with_suffix('.xlsx').unlink() for x in sht_2del]

  remove_excess_sheets(df)

  ##
  def remove_excess_rows(
      idf
      ) :
    print('len:' , len(idf))
    _by = return_index_level_names_from_to(0 , 2)
    _cou = count_notna_of_thecol_in_groups(idf , _by , indi.L3_SheetNum)
    rc = _cou['helper'].ge(1)
    rc &= idf[indi.L3_SheetNum].isna()
    idf = idf[~rc]
    print(len(idf))

    idf = idf.drop(columns = 'helper')
    return idf

  df = remove_excess_rows(df)

  ##
  def mask_excels(
      idf
      ) :
    _cnd = idf[cc.fileSuf].eq('.xlsx')
    _cnd &= idf[indi.L3_SheetNum].isna()
    print('len:' , len(_cnd[_cnd]))
    return _cnd

  msk = mask_excels(df)

  ##
  def process_all_excels(
      idf ,
      imsk
      ) :
    flt = idf[imsk]
    print("xl to review count:" , len(flt))

    cpc = mp.cpu_count()
    pool = mp.Pool(cpc)

    _res = pool.map(_targ , flt[cc.fileStem])

    return _res

  res = process_all_excels(df , msk)

  ##
  def _add2df(
      idf ,
      sheetname_hidestat_sheetname_isempty ,
      index
      ) :
    shsi = sheetname_hidestat_sheetname_isempty
    sh = shsi[0]
    si = shsi[1]

    _brow = idf.loc[index].to_frame().T.copy()

    sheet_cou = len(sh)
    not_hidden_kys = [ky for ky , va in sh.items() if not va]
    not_hidden_cou = len(not_hidden_kys)

    if sheet_cou > 0 :
      idf = idf.drop(index = index)
    for _i , (ky , val) in enumerate(sh.items()) :
      _nrow = _brow.copy()
      _nrow[cc.sheet_name] = str(ky)
      _nrow[cc.is_sheet_hidden] = val
      _nrow[cc.is_empty] = si[ky]
      _nrow[cc.sheet_count] = sheet_cou
      _nrow[indi.L3_SheetNum] = _i
      _nrow[cc.not_hidden_sheets_cou] = not_hidden_cou
      idf = pd.concat([idf , _nrow])
    return idf

  def add2df_all(
      idf ,
      imsk ,
      ires
      ) :
    _wanted_indices = imsk[imsk].index
    for _ind , _hs in zip(_wanted_indices , ires) :
      idf = _add2df(idf , _hs , _ind)
    return idf

  df = add2df_all(df , msk , res)

  ##
  df = remove_excess_rows(df)

  ##
  def build_sheet_stems(
      idf
      ) :
    _cnd = idf[indi.L3_SheetNum].notna()

    _the_col = idf.loc[_cnd , indi.L3_SheetNum]
    idf.loc[_cnd , indi.L3_SheetNum] = _the_col.astype(int).astype(str)
    _df = idf.loc[_cnd , [cc.fileStem , indi.L3_SheetNum]]
    _df = _df.fillna('N')
    _df = _df.applymap(str)
    idf.loc[_cnd , cc.sheetStem] = _df.agg('_'.join , axis = 1)
    return idf

  df = build_sheet_stems(df)

  ##
  save_current_module_cache(df , cur_mod_cache_fp)

  ##
  remove_excess_sheets(df)

  ##
  save_current_module_cache(df , cur_mod_cache_fp)

##


if __name__ == "__main__" :
  main()
  print(f'{Module_name} Done.')

def _test() :
  pass

  ##
