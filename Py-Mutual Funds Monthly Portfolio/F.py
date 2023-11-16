# %%
import pandas as pd
import main as m
from multiprocess import Pool
from multiprocessing import cpu_count
from zipfile import BadZipFile
from openpyxl import load_workbook
import os
import signal
import numpy as np


# %%
cur_n = 'F'
pre_n = 'E'

# %%
cur_prq = m.within_exe_data_dir + f'/{cur_n}' + m.prqsuf
pre_prq = m.within_exe_data_dir + f'/{pre_n}' + m.prqsuf

outputs = ['XlErr' , 'SheetNames' , 'XlSavedAsPrq']
outputs_dict = m.make_output_dict(outputs)
outputs_dict


# %%

class TimeoutError(Exception) :
    pass


def handler() :
    raise TimeoutError()


signal.signal(signal.SIGALRM , handler)


# %%


# noinspection PyTypeChecker
class xl_wb :


    def __init__(self , tracingno) :
        self.tracingno = str(tracingno)
        self.xlpn = m.excels_dir + '/' + str(tracingno) + '.xlsx'
        self.wb_dict = None
        self.outputs = outputs_dict.copy()


    def load_workbook_as_dict(self) :
        wb = load_workbook(self.xlpn , read_only = True , data_only = True , keep_links = False , keep_vba = False)

        if is_wb_loaded_by_openpy(wb) :
            workbook_data = {}
            for sh in wb.sheetnames :
                workbook_data[sh] = pd.DataFrame(wb[sh].values)
            return workbook_data

        workbook_data = pd.read_excel(self.xlpn , sheet_name = None)
        if is_wb_loaded_by_pandas(workbook_data) :
            return workbook_data


    def process_and_make_outputs(self) :
        signal.alarm(30)

        try :
            self.wb_dict = self.load_workbook_as_dict()

            self.outputs['SheetNames'] = {}
            for shno , key_df in enumerate(self.wb_dict.items()) :
                if not key_df[1].empty :
                    self.outputs['SheetNames'][shno] = key_df[0]
                    m.save_as_parquet(key_df[1] , f'{m.excels_as_prq_dir}/{self.tracingno}_{shno}{m.prqsuf}')

            self.outputs['XlErr'] = None
            self.outputs['XlSavedAsPrq'] = True
            self.outputs['SheetNames'] = str(self.outputs['SheetNames'])

            signal.alarm(0)
            print(self.outputs)
            return self.outputs
        except TimeoutError as e :
            print(e)
            self.outputs['XlErr'] = 'TimeOut'
            signal.alarm(0)
            return self.outputs


def is_wb_loaded_by_openpy(openpyxlwb) :
    for sh in openpyxlwb.sheetnames :
        sheet_data = pd.DataFrame(openpyxlwb[sh].values)
        if sheet_data.shape[0] > 4 and sheet_data.shape[1] > 4 :
            return True
    return False


def is_wb_loaded_by_pandas(wbdict) :
    for val in wbdict.values() :
        if val.shape[0] > 4 and val.shape[1] > 4 :
            return True
    return False


def target(tracingno) :
    obj = xl_wb(tracingno)
    out = obj.process_and_make_outputs()
    return list(out.values())


def main() :
    pass

    # %%
    df = m.read_parquet(pre_prq)
    df

    # %%
    df[outputs] = None

    # %%
    df = m.update_with_last_data(df , cur_prq)

    # %%
    df['IsXlDownloaded'] = df['TracingNo'].apply(lambda x : os.path.exists(f'{m.excels_dir}/{str(x)}.xlsx'))

    # %%
    cond = df['IsXlDownloaded'].eq(True)
    cond[cond]

    # %%
    cond &= df['XlSavedAsPrq'].isna() | df['XlErr'].eq('TimeOut')
    cond[cond]

    # %%
    flt = df[cond]
    flt

    # %%
    cores_n = cpu_count()
    print(f'Num of cores : {cores_n}')
    clstrsInd = m.return_clusters_indices(flt)
    pool = Pool(cores_n)

    # %%
    for i in range(0 , len(clstrsInd) - 1) :
        starti = clstrsInd[i]
        endi = clstrsInd[i + 1]
        print(f"Cluster index of {starti} to {endi}")

        indices = flt.iloc[starti :endi].index

        tracingnos = flt.loc[indices , 'TracingNo']

        out = pool.map(target , tracingnos)

        for j , outpa in enumerate(outputs) :
            df.loc[indices , outpa] = [w[j] for w in out]

        print(df.loc[indices , outputs])
        break

        # %%
    m.save_as_parquet(df , cur_prq)

    # %%
    cnd2 = df['XlErr'].isin(['TimeOut'])
    flt2 = df[cnd2]
    flt2

    # %%

#
# if __name__ == '__main__' :
#     main()
#
# # noinspection PyUnreachableCode
# if False :
#     pass
#
#     # %%
