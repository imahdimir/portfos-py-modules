"""[ ]

"""

# portfos env

##

import pathlib
import re

import openpyxl as pyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from persiantools import digits

# Important Params

mdir = pathlib.Path('/Users/mahdimir/Dropbox/R-Heidari-Mir/Energy_Pasargad')  # main dir
price_data_fp = '/Users/mahdimir/Documents/portfos_data/m_helper_data/Cleaned_Stock_Prices_1400_06_16.parquet'
tse_pr_fp = mdir / 'data' / 'tse-profits-97.xlsx'
tse_bs_fp = mdir / 'data' / 'tse-bs.xlsx'
tse_save_fp = mdir / 'data' / 'tse300.xlsx'  # first 300 tse firms fin data save fp

tse_year = 1397

def find_jdate_as_int_using_re(
    ist
    ) :
  ist = digits.fa_to_en(str(ist))
  match_obj = re.search(r'1[3-4]\d{2}/?[0-1]\d/?[0-3]\d' , ist)
  if match_obj :
    match = match_obj.group()
    num = re.sub(r'\D' , '' , match)
    return str(int(num))

def find_n_jmonth_ahead(
    cur_month ,
    howmany=1
    ) :
  if howmany == 1 :
    if cur_month % 100 == 12 :
      next_month = (cur_month // 100 + 1) * 100 + 1
    else :
      next_month = cur_month + 1
    return next_month

  else :
    next_month = find_n_jmonth_ahead(cur_month)
    return find_n_jmonth_ahead(next_month , howmany - 1)

def save_df_as_a_nice_xl(
    df: pd.DataFrame ,
    fpn: pathlib.Path = None ,
    index: bool = False ,
    header: bool = True ,
    max_col_length: int = 100
    ) :
  wb = pyxl.Workbook()
  ws = wb.active
  df = df.fillna(value='')
  for r in dataframe_to_rows(df , index=index , header=header) :
    ws.append(r)
  panes = index * ws['A'] + header * ws[1]
  for cell in panes :
    cell.style = 'Pandas'
  for column in ws.columns :
    length = max(len(str(cell.value)) for cell in column) + 3
    length = length if length <= max_col_length else max_col_length
    ws.column_dimensions[column[0].column_letter].width = length
  ws.freeze_panes = 'A2'
  if fpn is not None :
    wb.save(fpn)
  wb.close()
  return wb

def read_tse_proftis_return_f300_data() :
  """
  reads tse profit and loss statements data, filters for needed cols,
  keep the data only for the wanted year then keeps only symbols that are
  present in if300 list.
  """

  ##
  # local params
  first_n = 300

  ##
  pro = pd.read_excel(tse_pr_fp , engine='openpyxl')

  ##
  pro['fiscal_year'] = pro['سال مالی'].apply(find_jdate_as_int_using_re)

  colspr = {
      'symbol'        : 'symbol' ,
      'fiscal_year'   : 'fiscal_year' ,
      'mktcap'        : 'ارزش روز' ,
      'op_rev'        : 'جمع درآمدها' ,
      'gross_profits' : 'سود ناویژه' ,
      'op_profits'    : 'سود (زیان) عملیاتی' ,
      'fiscal_costs'  : 'هزینه‌های مالی' ,
      'net_profits'   : 'سود (زیان) پس از کسر مالیات' ,
      }

  cols2 = colspr.values()
  pro = pro[cols2]

  pro = pro.rename(columns={y : x for x , y in colspr.items()})

  msk = pro['fiscal_year'].notna()
  pro1 = pro[msk]

  ##
  msk = pro1['fiscal_year'].astype(int).le(int(str(tse_year + 1) + '0431'))
  msk &= pro1['fiscal_year'].astype(int).ge(int(str(tse_year) + '0930'))

  pro2 = pro1[msk]

  ##
  pro2 = pro2.sort_values(by='mktcap' , ascending=False)

  ##
  Pro = pro2.iloc[:first_n]
  F300 = Pro['symbol']

##

def return_biggest_firms_balancesheet_data() :
  """

  """

  ##
  ase = 'approved_shareholders_equity'
  she_col = 'shareholders_equity'
  fyr_c = 'fiscal_year'
  sym = 'symbol'
  asset = 'total_assets'

  init_asset = 'init_' + asset
  init_ase = 'init_' + ase

  ##
  bs = pd.read_excel(tse_bs_fp , engine='openpyxl')

  ##
  msk = bs[ase].isna()
  bs.loc[msk , ase] = bs[she_col]

  ##
  bs[fyr_c] = bs[fyr_c].astype(int).astype(str)
  bs = bs.sort_values(by=fyr_c)

  ##
  bs['jm'] = bs[fyr_c].astype(int) // 100
  bs['1_yr'] = bs['jm'].apply(lambda
                                x : find_n_jmonth_ahead(x , 12))

  ##
  bs1 = bs[[sym , '1_yr' , asset , ase]]

  bs1 = bs1.rename(columns={
      asset : init_asset ,
      ase   : init_ase ,
      })

  bs1['1_yr'] = bs1['1_yr'].astype(int).astype(str)
  bs['jm'] = bs['jm'].astype(int).astype(str)

  ##
  bs = bs.drop(columns=['1_yr'])

  ##
  Bs = bs.merge(bs1 ,
                left_on=[sym , 'jm'] ,
                right_on=[sym , '1_yr'] ,
                how='left')

  ##
  cl = 'current_liabilities'
  ncl = 'noncurrent_liabilities'
  cua = 'current_assets'
  ncua = 'noncurrent_assets'
  nfa = 'net_fixed_assets'

  cols = {
      sym        : None ,
      fyr_c      : None ,
      cl         : None ,
      ncl        : None ,
      ase        : None ,
      cua        : None ,
      ncua       : None ,
      nfa        : None ,
      init_asset : None ,
      init_ase   : None ,
      }

  Bs = Bs[cols.keys()]

##

def combine_profit_and_balancesheet() :
  """

  """

  ##
  Tse = Pro.merge(Bs , on=[sym , fyr_c] , how='left')

  ##
  save_df_as_a_nice_xl(Tse , tse_save_fp)

  ##
