"""[ ]

"""

##

import pathlib

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import pandas as pd
import seaborn as sns
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from matplotlib.ticker import FuncFormatter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats

mdir = pathlib.Path('/Users/mahdimir/Dropbox/R-Heidari-Mir/Energy_Pasargad')  # main dir

tse_ratios_fp = mdir / 'data' / 'tse300-ratios.xlsx'
tse_desc_fp = mdir / 'data' / 'tse300-desc.xlsx'

def save_df_as_a_nice_xl(
    df: pd.DataFrame ,
    fpn: pathlib.Path = None ,
    index: bool = False ,
    header: bool = True ,
    max_col_length: int = 100
    ) :
  wb = pyxl.Workbook()
  ws = wb.active
  df = df.fillna(value='')
  for r in dataframe_to_rows(df , index=index , header=header) :
    ws.append(r)
  panes = index * ws['A'] + header * ws[1]
  for cell in panes :
    cell.style = 'Pandas'
  for column in ws.columns :
    length = max(len(str(cell.value)) for cell in column) + 3
    length = length if length <= max_col_length else max_col_length
    ws.column_dimensions[column[0].column_letter].width = length
  ws.freeze_panes = 'A2'
  if fpn is not None :
    wb.save(fpn)
  wb.close()
  return wb

def add_ratios(
    df , ratios_dict
    ) :
  for ky , val in ratios_dict.items() :
    msk = df[val[1]].ne(0)
    df.loc[msk , ky] = df.loc[msk , val[0]] / df.loc[msk , val[1]]
  return df

def filter_nan_and_outliers(
    iser: pd.Series , param=3
    ) :
  iser = iser[iser.notna()]

  z_scores = stats.zscore(iser)
  abs_z_scores = np.abs(z_scores)
  filter_outliers = abs_z_scores < param
  iser = iser[filter_outliers]

  return iser

gost = 'گسترش انرژی پاسارگاد'
gost = get_display(reshape(gost))

miane = get_display(reshape('میانه'))

##
def read_data() :
  """

  """

  ##
  Tse = pd.read_excel(tse_ratios_fp , engine='openpyxl')
  Desc = pd.read_excel(tse_desc_fp , engine='openpyxl')

##

def op_profits_margin_dist() :
  """

  """

  ##
  opmrg = 'op_profits_margin'

  ##
  x = Tse[opmrg]
  x = filter_nan_and_outliers(x)
  x = x * 100  # pct

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True)
  ax.set_xlim(-120 , 200)

  ##
  # ax.set_xlabel('Operational Profits Margin')
  ax.set_xlabel("")
  ax.xaxis.set_major_formatter(FuncFormatter(lambda a , _ : "{:g}%".format(a)))

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 97
  plt.axvline(gostval , linestyle='-' , label=gost , color='g')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      sadak = _i
      break
  print(x.quantile(sadak / 100))

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400)

##

def net_profits_margin_dist() :
  """

  """

  ##
  npmrg = 'net_profits_margin'

  ##
  x = Tse[npmrg]
  x = filter_nan_and_outliers(x)
  x = x * 100  # pct

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True , color='orange')
  ax.set_xlim(-120 , 200)

  ##
  ax.set_xlabel('Net Profits Margin')
  ax.xaxis.set_major_formatter(FuncFormatter(lambda a , _ : "{:g}%".format(a)))

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 4
  plt.axvline(gostval , linestyle='-' , label=gost , color='b')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      break
  print(x.quantile())

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400 , transparent=True)

##
def current_ratio_dist() :
  """

  """

  ##
  curr = 'current_ratio'

  ##
  x = Tse[curr]
  x = filter_nan_and_outliers(x)

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True , color='r' , cut=True)
  ax.set_xlim(-1 , 4)

  ##
  ax.set_xlabel('Current Ratio')

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 0.76
  plt.axvline(gostval , linestyle='-' , label=gost , color='b')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      break
  print(x.quantile())

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400 , transparent=True)

##

def debt_ratio_dist() :
  """

  """

  ##
  debr = 'debt_ratio'

  ##
  x = Tse[debr]
  x = filter_nan_and_outliers(x)

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True , color='g' , cut=True)

  ##
  ax.set_xlabel('Debt Ratio')

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 0.95
  plt.axvline(gostval , linestyle='-' , label=gost , color='b')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      break
  print(x.quantile())

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400 , transparent=True)

##

def roa_dist() :
  """

  """

  ##
  roa = 'ROA'

  ##
  x = Tse[roa]
  x = filter_nan_and_outliers(x)
  x = x * 100

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True , color='m' , cut=True)

  ##
  ax.set_xlabel('ROA')
  ax.xaxis.set_major_formatter(FuncFormatter(lambda a , _ : "{:g}%".format(a)))

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 1
  plt.axvline(gostval , linestyle='-' , label=gost , color='b')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      break
  print(x.quantile())

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400 , transparent=True)

  ##

def roe_dist() :
  """

  """

  ##
  roe = 'ROE'

  ##
  x = Tse[roe]
  x = filter_nan_and_outliers(x)
  x = x * 100

  ##
  plt.close()
  sns.set_style('white')
  ax = sns.kdeplot(x , fill=True , color='y' , cut=True)
  ax.set_xlim(-100 , 220)

  ##
  ax.set_xlabel('ROE')
  ax.xaxis.set_major_formatter(FuncFormatter(lambda a , _ : "{:g}%".format(a)))

  ax.yaxis.set_ticks([])
  ax.set_ylabel('')

  ##
  plt.axvline(x.median() , linestyle='-' , label=miane , color='r')
  print(x.median())
  print(x.quantile(.1))
  print(x.quantile(.9))

  ##
  gostval = 15.4
  plt.axvline(gostval , linestyle='-' , label=gost , color='b')
  for _i in range(0 , 101) :
    if gostval < x.quantile(_i / 100) :
      print(_i)
      break
  print(x.quantile())

  ##
  font = {
      'family' : 'B Mitra' ,
      'size'   : 12
      }
  matplotlib.rc('font' , **font)

  ##
  ax.legend()

  ##
  plt.savefig('1.png' , dpi=400 , transparent=True)

##
def market_caps() :
  """

  """

  ##
  prices_fp = '/Users/mahdimir/Documents/portfos_data/m_helper_data/Cleaned_Stock_Prices_1400_06_16.parquet'
  prdf = pd.read_parquet(prices_fp)

  ##
  msk = prdf['name'].str.contains('فار')
  df1 = prdf[msk]

  ##

  ##

  oprev = 'op_rev'
  oppro = 'op_profits'
  gpro = 'gross_profits'
  npro = 'net_profits'

  gpromrg = 'gross_profits_margin'
  npmrg = 'net_profits_margin'
  curr = 'current_ratio'
  debr = 'debt_ratio'
  roa = 'ROA'
  roe = 'ROE'

  ratios = {
      opmrg   : (oppro , oprev) ,
      gpromrg : (gpro , oprev) ,
      npmrg   : (npro , oprev) ,
      curr    : (cua , cl) ,
      debr    : (tol , tos) ,
      roa     : (npro , ava) ,
      roe     : (npro , avse) ,
      }

  ##
  Tse = add_ratios(Tse , ratios)

##

def desc_tse300_save() :
  """

  """

  ##
  save_df_as_a_nice_xl(Tse , tse_ratios_fp)

  ##
  desc = Tse.describe()

  ##
  save_df_as_a_nice_xl(desc , tse_desc_fp , index=True)
