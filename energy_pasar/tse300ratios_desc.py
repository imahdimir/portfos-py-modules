"""[ ]

"""

##

import pathlib

import openpyxl as pyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

mdir = pathlib.Path('/Users/mahdimir/Dropbox/R-Heidari-Mir/Energy_Pasargad')  # main dir
tse_fp = mdir / 'data' / 'tse300.xlsx'  # first 300 tse firms fin data save fp

tse_ratios_fp = mdir / 'data' / 'tse300-ratios.xlsx'
tse_desc_fp = mdir / 'data' / 'tse300-desc.xlsx'

def save_df_as_a_nice_xl(
    df: pd.DataFrame ,
    fpn: pathlib.Path = None ,
    index: bool = False ,
    header: bool = True ,
    max_col_length: int = 100
    ) :
  wb = pyxl.Workbook()
  ws = wb.active
  df = df.fillna(value='')
  for r in dataframe_to_rows(df , index=index , header=header) :
    ws.append(r)
  panes = index * ws['A'] + header * ws[1]
  for cell in panes :
    cell.style = 'Pandas'
  for column in ws.columns :
    length = max(len(str(cell.value)) for cell in column) + 3
    length = length if length <= max_col_length else max_col_length
    ws.column_dimensions[column[0].column_letter].width = length
  ws.freeze_panes = 'A2'
  if fpn is not None :
    wb.save(fpn)
  wb.close()
  return wb

def add_ratios(
    df ,
    ratios_dict
    ) :
  for ky , val in ratios_dict.items() :
    msk = df[val[1]].ne(0)
    df.loc[msk , ky] = df.loc[msk , val[0]] / df.loc[msk , val[1]]
  return df

##
def define_agg_cols_for_tse_300() :
  """

  """

  ##
  cua = 'current_assets'
  ncua = 'noncurrent_assets'
  tos = 'total_assets'
  itos = 'init_' + tos
  ava = 'avg_assets'

  ase = 'approved_shareholders_equity'
  iase = 'init_' + ase
  avse = 'avg_shareholders_equity'

  cl = 'current_liabilities'
  ncl = 'noncurrent_liabilities'
  tol = 'total_liabilities'

  ##
  tse = pd.read_excel(tse_fp , engine='openpyxl')

  ##
  tse[tos] = tse[cua] + tse[ncua]
  tse[ava] = tse[[itos , tos]].mean(axis=1)
  tse[avse] = tse[[iase , ase]].mean(axis=1)
  tse[tol] = tse[cl] + tse[ncl]

  ##
  Tse = tse.copy()

##

def cal_ratios() :
  """

  """

  ##
  oprev = 'op_rev'
  oppro = 'op_profits'
  gpro = 'gross_profits'
  npro = 'net_profits'

  opmrg = 'op_profits_margin'
  gpromrg = 'gross_profits_margin'
  npmrg = 'net_profits_margin'
  curr = 'current_ratio'
  debr = 'debt_ratio'
  roa = 'ROA'
  roe = 'ROE'

  ratios = {
      opmrg   : (oppro , oprev) ,
      gpromrg : (gpro , oprev) ,
      npmrg   : (npro , oprev) ,
      curr    : (cua , cl) ,
      debr    : (tol , tos) ,
      roa     : (npro , ava) ,
      roe     : (npro , avse) ,
      }

  ##
  Tse = add_ratios(Tse , ratios)

##

def desc_tse300_save() :
  """

  """

  ##
  save_df_as_a_nice_xl(Tse , tse_ratios_fp)

  ##
  desc = Tse.describe()

  ##
  save_df_as_a_nice_xl(desc , tse_desc_fp , index=True)
